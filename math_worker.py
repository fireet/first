#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct  1 12:55:14 2019

@author: fireet
"""
from data_base_maker import get_data, get_complex_tars
import openpyxl
from openpyxl.styles.borders import Border, Side
import os
import time

start_time = time.time()

thin_border = Border(bottom=Side(style='thin'))


class Complex:
    ind = {}

    def __init__(self,
                 complex_name='',
                 hot_water_tar=0.0,
                 cold_water_tar=0.0,
                 electro_tar=0.0,
                 water_off_tar=0.0,
                 ):

        self.complex_name = complex_name
        self.hot_water_tar = hot_water_tar
        self.cold_water_tar = cold_water_tar
        self.water_off_tar = water_off_tar
        self.electro_tar = electro_tar
        self.arend = {}
        Complex.ind[complex_name] = self

    def add_arend(self, doc, name):
        tars = dict(ГВС=self.hot_water_tar,
                    ХВС=self.cold_water_tar,
                    ЭЭ=self.electro_tar,
                    ОТВОД=self.water_off_tar)
        self.arend[str(doc)] = (Arend(doc=str(doc),
                                      name=name,
                                      tars=tars))


class Res:
    def __init__(self,
                 name='',
                 old=0.0,
                 new=0.0,
                 coef=1.0,
                 tars={}):

        self.name = name.split()[0].upper()
        self.old = old
        self.new = new
        self.total = 0.0
        self.tars = tars[self.name]
        self.coef = coef
        self.cash = 0.0

    def work_res(self):
        self.total = round(self.new - self.old, 2)
        self.cash = round(self.total*self.coef*self.tars*self.coef, 2)


class Arend:
    arend_pool = []

    def __init__(self,
                 doc='',
                 name='',
                 tars={}):

        self.doc = str(doc)
        self.name = name
        self.total_cash = 0.0
        self.data_res = []
        self.tars = tars
        self.lost = 0.0
        Arend.arend_pool.append(self)

    def new_res(self,
                name_res='',
                old=0.0,
                new=0.0,
                coef=1.0):

        self.data_res.append(Res(name=name_res,
                                 old=old,
                                 new=new,
                                 tars=self.tars,
                                 coef=coef))

    def get_all_value(self):
        pool = {}
        total_cash_res = 0
        for i in self.data_res:
            if i.name not in pool:
                pool[i.name] = 0

            pool[i.name] += round(i.cash, 2)
            total_cash_res += i.cash
            self.total_cash += round(i.cash, 2)
        self.data_res.append(pool)


names = get_complex_tars()

Complex(**names['первый'])
active = Complex.ind['первый']

data = get_data()

for item in data:
    if item[0] not in active.arend and item[0]:
        doc = str(item[0])
        name = item[1]
        active.add_arend(doc, name)
    tmp = dict(name_res=item[2],
               old=item[3],
               new=item[4],
               coef=item[5],
               )
    active.arend[str(doc)].new_res(**tmp)
    for i in active.arend[str(doc)].data_res:
        i.work_res()
for x in active.arend.values():
    x.get_all_value()


if not os.path.isfile('test.xlsx'):
    wb = openpyxl.Workbook()
    wb.save('test.xlsx')

for x in Complex.ind.values():
    wbk = openpyxl.load_workbook('test.xlsx')
    wbs = wbk[wbk.sheetnames[0]]
    row = 1
    colomn = 1
    wbs.cell(row, colomn).value = x.complex_name
    wbk.save('test.xlsx')
    row += 1
    for s in x.arend.values():

        for i in ['doc', 'name', 'data_res', 'total_cash']:
            if i == 'data_res':
                s.__dict__[i].pop()
                pool = s.__dict__[i]
                for item in pool:
                    column = colomn
                    for path in item.__dict__.keys():
                        tmp = item.__dict__[path]
                        if type(tmp) != str:
                            tmp = '{:0.2f}'.format(tmp)
                        wbs.cell(row, column).value = item.__dict__[path]
                        wbk.save('test.xlsx')
                        column += 1
                    row += 1
                colomn = column
            else:
                tmp = s.__dict__[i]
                if type(tmp) is not str:
                    tmp = '{:0.2f}'.format(tmp)
                wbs.cell(row, colomn).value = str(tmp)
                colomn += 1
                wbk.save('test.xlsx')
        for i in range(1, wbs.max_column+1):
            wbs.cell(row, i).border = thin_border
        colomn = 1
        row += 1
        wbk.save('test.xlsx')
    wbk.close
    break


print('--- %s seconds ---' % (time.time() - start_time))
 