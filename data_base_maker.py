#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep  4 13:40:04 2019

@author: fireet
"""

import openpyxl


def open_xlsx(file_name):
    wb = openpyxl.load_workbook(file_name)
    work = wb.active
    return work


def get_complex_tars():
    work = open_xlsx('./data/complex_tars.xlsx')
    data = {}
    for row in range(2, work.max_row+1):
        complex_name = work.cell(row, 1).value
        item = dict(complex_name=complex_name,
                    hot_water_tar=work.cell(row, 2).value,
                    cold_water_tar=work.cell(row, 3).value,
                    water_off_tar=work.cell(row, 4).value,
                    electro_tar=work.cell(row, 5).value,
                    )
        data[complex_name] = item
    return data


def get_data():
    work = open_xlsx('./data/59322001.xlsx')
    data = []
    for row in range(4, work.max_row+1):
        tmp = []
        for column in range(1, 7):
            tmp.append(work.cell(row, column).value)
        data.append(tmp)

    return data
